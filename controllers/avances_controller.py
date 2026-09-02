from flask import Blueprint, render_template, request, redirect, url_for, flash, session
from models import db, Actividades, Avances, Proyectos, AvanceMaterial, Evidencias, Usuarios, Notificaciones, Materiales
from datetime import datetime
import os, base64
from werkzeug.utils import secure_filename
import uuid
import requests
from supabase_client import supabase
from frases import obtener_frase
from decorators import login_required, admin_required, admin_encargado_required
from flask import send_file
from openpyxl import load_workbook
from io import BytesIO
from openpyxl.drawing.image import Image as ExcelImage


# Carpeta donde se guardarán las imágenes
UPLOAD_FOLDER = "static/uploads/evidencias"
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif"}

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def parse_float(val):
    if val is None or str(val).strip() == "":
        return None
    try:
        return float(val)
    except ValueError:
        return None

avances_bp = Blueprint("avances", __name__)


# ===============================================================
# REGISTRAR AVANCE DE UNA ACTIVIDAD (trabajador)
# ===============================================================
@avances_bp.route("/registrar/<int:id_actividad>", methods=["POST"])
@login_required
def registrar_avance(id_actividad):
    flash(obtener_frase("avance"), "success")
    id_usuario = session.get("user_id")

    if not id_usuario:
        flash("Debes iniciar sesión para enviar avances.", "warning")
        return redirect(url_for("auth.login"))

    try:
        # ==========================
        # Datos base
        # ==========================
        unidades = int(request.form["unidades_avanzadas"])
        fecha_str = request.form["fecha"]
        fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        mensaje = request.form.get("mensaje", "")

        # Campos de ubicación...
        trayecto = request.form.get("trayecto")
        calzada = request.form.get("calzada")
        carril = request.form.get("carril")
        ubicacion_pr = request.form.get("ubicacion_pr")
        tipo = request.form.get("tipo")
        elemento = request.form.get("elemento")
        area_elemento = parse_float(request.form.get("area_elemento"))
        area_total = parse_float(request.form.get("area_total"))
        
        # Nuevos campos de ubicación/geometría
        margen = request.form.get("margen")
        pr_inicio = request.form.get("pr_inicio")
        pr_fin = request.form.get("pr_fin")
        longitud_lineal = parse_float(request.form.get("longitud_lineal"))
        color_lineal = request.form.get("color_lineal")
        ancho = parse_float(request.form.get("ancho"))
        largo = parse_float(request.form.get("largo"))
        cantidad = parse_float(request.form.get("cantidad"))
        tamano = request.form.get("tamano")
        color = request.form.get("color")

        # ==========================
        # Crear avance
        # ==========================
        nuevo_avance = Avances(
            id_actividad=id_actividad,
            id_usuario=id_usuario,
            fecha=fecha,
            unidades_avanzadas=unidades,
            mensaje=mensaje,
            trayecto=trayecto,
            calzada=calzada,
            carril=carril,
            ubicacion_pr=ubicacion_pr,
            tipo=tipo,
            elemento=elemento,
            area_elemento=area_elemento,
            area_total=area_total,
            margen=margen,
            pr_inicio=pr_inicio,
            pr_fin=pr_fin,
            longitud_lineal=longitud_lineal,
            color_lineal=color_lineal,
            ancho=ancho,
            largo=largo,
            cantidad=cantidad,
            tamano=tamano,
            color=color,
        )

        db.session.add(nuevo_avance)
        # 🟢 NUEVO: Usamos flush para obtener el ID del avance sin cerrar la transacción
        db.session.flush()

        # ==================================================
        # 🟢 NUEVO: PROCESAR MATERIALES USADOS (ESTUDIO EFICIENCIA)
        # ==================================================
        # Buscamos en el formulario campos que empiecen con 'material_'
        for key in request.form:
            if key.startswith("material_"):
                try:
                    # Extraemos el ID del material del nombre del input (ej: material_5 -> 5)
                    id_material = int(key.split("_")[1])
                    cantidad_usada = float(request.form.get(key) or 0)

                    if cantidad_usada > 0:
                        # 1. Registrar el consumo para el estudio comparativo
                        consumo = AvanceMaterial(
                            id_avance=nuevo_avance.id_avance,
                            id_material=id_material,
                            cantidad_usada=cantidad_usada
                        )
                        db.session.add(consumo)

                        # 2. Descontar del stock global en BODEGA (Tabla Materiales)
                        mat_inventario = Materiales.query.get(id_material)
                        if mat_inventario:
                            mat_inventario.cantidad -= cantidad_usada
                except (ValueError, IndexError):
                    continue # Si hay un error con un campo, sigue con el siguiente



        # ==================================================
        # Guardar evidencias (galería o cámara) en Supabase
        # ==================================================
        files = request.files.getlist("evidencias")

        # 1️⃣ Archivos desde galería
        for file in files:
            if file and allowed_file(file.filename):
                original_filename = secure_filename(file.filename)
                ext = original_filename.rsplit('.', 1)[1].lower() if '.' in original_filename else 'jpg'
                filename = f"{uuid.uuid4().hex}.{ext}"
                
                # Leer el archivo en memoria
                file_bytes = file.read()
                
                if supabase:
                    # Subir a Supabase
                    supabase.storage.from_("evidencias").upload(
                        path=filename, 
                        file=file_bytes, 
                        file_options={"content-type": file.content_type}
                    )
                    # Obtener URL pública
                    public_url = supabase.storage.from_("evidencias").get_public_url(filename)
                else:
                    # Fallback local por si acaso no hay keys
                    ruta_relativa = os.path.join("uploads", "evidencias", filename)
                    ruta_completa = os.path.join("static", ruta_relativa)
                    os.makedirs(os.path.dirname(ruta_completa), exist_ok=True)
                    with open(ruta_completa, "wb") as f:
                        f.write(file_bytes)
                    public_url = ruta_relativa

                evidencia = Evidencias(
                    id_avance=nuevo_avance.id_avance,
                    ruta_archivo=public_url,
                    tipo="imagen"
                )
                db.session.add(evidencia)

        # Imagen tomada con cámara (base64)
        imagen_capturada = request.form.get("captura_base64")
        if imagen_capturada:
            img_data = base64.b64decode(imagen_capturada.split(",")[1])
            filename = f"captura_{uuid.uuid4().hex}.jpg"
            
            if supabase:
                # Subir a Supabase
                supabase.storage.from_("evidencias").upload(
                    path=filename, 
                    file=img_data, 
                    file_options={"content-type": "image/jpeg"}
                )
                public_url = supabase.storage.from_("evidencias").get_public_url(filename)
            else:
                ruta_relativa = os.path.join("uploads", "evidencias", filename)
                ruta_completa = os.path.join("static", ruta_relativa)
                os.makedirs(os.path.dirname(ruta_completa), exist_ok=True)
                with open(ruta_completa, "wb") as f:
                    f.write(img_data)
                public_url = ruta_relativa

            evidencia = Evidencias(
                id_avance=nuevo_avance.id_avance,
                ruta_archivo=public_url,
                tipo="imagen"
            )
            db.session.add(evidencia)

        # ==================================================
        # Notificar a ADMIN
        # ==================================================
        admins = Usuarios.query.filter_by(rol="ADMIN").all()

        if admins:
            actividad = Actividades.query.get(id_actividad)
            proyecto = actividad.proyecto

            mensaje_notif = (
                f"Se ha registrado un nuevo avance en la actividad "
                f"'{actividad.nombre}' del proyecto '{proyecto.nombre}'."
            )

            for admin in admins:
                notificacion = Notificaciones(
                    id_usuario_destino=admin.id_usuario,
                    mensaje=mensaje_notif
                )
                db.session.add(notificacion)

        # ==================================================
        # Recalcular progreso de la actividad
        # ==================================================
        actividad = Actividades.query.get(id_actividad)
        total = actividad.unidades_totales or 0

        avanzado = (
            db.session.query(db.func.sum(Avances.unidades_avanzadas))
            .filter_by(id_actividad=id_actividad)
            .scalar()
        ) or 0

        session[f'avance_{id_actividad}'] = {
            'avanzado': avanzado,
            'porcentaje': int((avanzado / total) * 100) if total > 0 else 0
        }

        db.session.commit() # 👈 Finaliza todo: avance, materiales e inventario
        flash("Avance y consumo de materiales registrados.", "success")

    except Exception as e:
        db.session.rollback()
        print("Error al registrar avance:", e)
        flash(f"Error: {str(e)}", "danger")

    return redirect(url_for("dashboard.dashboard_trabajador") + f"#actividad-{id_actividad}")


# ===============================================================
# INFORME DE AVANCE DE UN PROYECTO
# ===============================================================
@avances_bp.route("/informe/<int:id_proyecto>")
@login_required
@admin_required
def ver_informe_avance(id_proyecto):
    id_usuario = session.get("user_id")

    if not id_usuario:
        flash("Debes iniciar sesión para ver el informe.", "warning")
        return redirect(url_for("auth.login"))

    proyecto = Proyectos.query.get_or_404(id_proyecto)

    avances = (
        db.session.query(Avances, Actividades)
        .join(Actividades, Actividades.id_actividad == Avances.id_actividad)
        .filter(Actividades.id_proyecto == id_proyecto)
        .order_by(Avances.fecha.desc())
        .all()
    )

    return render_template(
        "informe_avance.html",
        proyecto=proyecto,
        avances=avances
    )


# ===============================================================
# Marcar notificación como leída
# ===============================================================
@avances_bp.route('/notificacion/leer/<int:id_notificacion>')
@login_required
def leer_notificacion(id_notificacion):
    notificacion = Notificaciones.query.get_or_404(id_notificacion)

    if notificacion.id_usuario_destino != session.get("user_id"):
        flash("No tienes permiso para ver esta notificación", "danger")
        return redirect(url_for('proyectos.manage_proyectos'))

    if not notificacion.leido:
        notificacion.leido = True
        db.session.commit()

    return redirect(url_for('proyectos.manage_proyectos'))


# ===============================================================
# Exportar informe de avance a Excel
# ===============================================================
@avances_bp.route("/informe/<int:id_proyecto>/excel")
@login_required
@admin_required
def exportar_informe_excel(id_proyecto):

    proyecto = Proyectos.query.get_or_404(id_proyecto)

    avances = (
        db.session.query(Avances, Actividades, Usuarios)
        .join(Actividades, Actividades.id_actividad == Avances.id_actividad)
        .join(Usuarios, Usuarios.id_usuario == Avances.id_usuario)
        .filter(Actividades.id_proyecto == id_proyecto)
        .order_by(Avances.fecha)
        .all()
    )

    # 📄 Abrir plantilla
    ruta_plantilla = os.path.join(
        "static", "templates_excel", "informe_avance.xlsx"
        )
    wb = load_workbook(ruta_plantilla)
    ws = wb.active

    # 🧾 Datos generales (ejemplo)
    ws["B2"] = proyecto.nombre
    ws["B3"] = proyecto.descripcion or ""

    fila = 3  # 👈 empieza debajo del encabezado

    for avance, actividad, usuario in avances:
        ws[f"A{fila}"] = avance.fecha.strftime("%d/%m/%Y") if avance.fecha else ""
        ws[f"B{fila}"] = avance.trayecto or ""
        ws[f"C{fila}"] = avance.calzada or ""
        ws[f"D{fila}"] = avance.carril or ""
        ws[f"E{fila}"] = avance.ubicacion_pr or ""
        ws[f"F{fila}"] = avance.tipo or ""
        ws[f"G{fila}"] = avance.elemento or ""
        ws[f"H{fila}"] = avance.unidades_avanzadas or ""
        ws[f"I{fila}"] = avance.area_elemento or ""
        ws[f"J{fila}"] = avance.area_total or ""
        
        # 📷 INSERTAR IMÁGENES (columna K)
        if avance.evidencias:
            for evidencia in avance.evidencias:
                if evidencia.ruta_archivo.startswith('http'):
                    # Es URL de Supabase, descargar en memoria
                    try:
                        response = requests.get(evidencia.ruta_archivo)
                        if response.status_code == 200:
                            img_stream = BytesIO(response.content)
                            img = ExcelImage(img_stream)
                            
                            img.width = 120
                            img.height = 90
                            celda_img = f"K{fila}"
                            ws.add_image(img, celda_img)
                            ws.row_dimensions[fila].height = 75
                            break
                    except Exception as e:
                        print("Error descargando imagen de Supabase para Excel:", e)
                else:
                    # Archivo local por retrocompatibilidad
                    ruta_imagen = os.path.join("static", evidencia.ruta_archivo)
                    if os.path.exists(ruta_imagen):
                        img = ExcelImage(ruta_imagen)
                        img.width = 120
                        img.height = 90
                        celda_img = f"K{fila}"
                        ws.add_image(img, celda_img)
                        ws.row_dimensions[fila].height = 75
                        break

        fila += 1


    # 💾 Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"informe_avance_{proyecto.nombre}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@avances_bp.route("/analisis_comparativo/<int:id_proyecto>")
@login_required
@admin_required
def analisis_comparativo(id_proyecto):
    proyecto = Proyectos.query.get_or_404(id_proyecto)
    
    # 1. Obtenemos lo planeado (de la tabla MaterialesProyecto)
    planeado = {mp.id_material: mp.cantidad for mp in proyecto.materiales}
    
    # 2. Obtenemos lo gastado (Sumando todos los avances de este proyecto)
    gastado_query = db.session.query(
        AvanceMaterial.id_material,
        db.func.sum(AvanceMaterial.cantidad_usada).label('total_usado')
    ).join(Avances).join(Actividades).filter(Actividades.id_proyecto == id_proyecto).group_by(AvanceMaterial.id_material).all()
    
    gastado = {item.id_material: item.total_usado for item in gastado_query}

    # 3. Consolidamos los datos para la tabla
    comparativa = []
    for mp in proyecto.materiales:
        id_m = mp.id_material
        uso_real = gastado.get(id_m, 0)
        plan = mp.cantidad
        diferencia = plan - uso_real
        porcentaje_uso = (uso_real / plan * 100) if plan > 0 else 0
        
        comparativa.append({
            'nombre': mp.material.nombre,
            'unidad': mp.material.unidad,
            'planeado': plan,
            'real': uso_real,
            'diferencia': diferencia,
            'porcentaje': porcentaje_uso
        })

    return render_template("analisis_comparativo.html", proyecto=proyecto, comparativa=comparativa)

@avances_bp.route("/seleccion_analisis")
@login_required
@admin_required
def seleccion_analisis():
    # Cambiamos Proyectos.creado_en por Proyectos.fecha_inicio
    proyectos = Proyectos.query.filter_by(visible=True).order_by(Proyectos.fecha_inicio.desc()).all()
    return render_template("seleccion_analisis.html", proyectos=proyectos)


# ===============================================================
# HISTORIAL DE AVANCES POR ACTIVIDAD (Trabajador)
# ===============================================================
@avances_bp.route("/historial/<int:id_actividad>")
@login_required
def historial_actividad(id_actividad):
    id_usuario = session.get("user_id")
    
    if not id_usuario:
        flash("Debes iniciar sesión para ver el historial.", "warning")
        return redirect(url_for("auth.login"))

    actividad = Actividades.query.get_or_404(id_actividad)
    
    # Obtener historial solo del trabajador actual para esta actividad
    avances = (
        db.session.query(Avances)
        .filter_by(id_actividad=id_actividad, id_usuario=id_usuario)
        .order_by(Avances.fecha.desc())
        .all()
    )
    
    return render_template("historial_avances.html", actividad=actividad, avances=avances)